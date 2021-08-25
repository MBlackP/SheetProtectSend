import getpass
from tkinter import *
from tkinter import filedialog
import openpyxl as xl
from openpyxl import Workbook

# steps:
# 1. get excel file name/path
# 2. get name of worksheet (month year)
# 3. create new excel file
# 4. get worksheet using worksheet name and copy it onto new excel file
# 5. name new excel file's worksheet
# 6. protect new excel file worksheet.
# 7. new excel file
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

root = Tk()
root.title('Securify')
root.geometry('700x500')
root.config(background = 'black')

def browseFile():

    global filePath
    filePath = filedialog.askopenfilename(initialdir = 'C:/Users/{}/', title = 'Select an Excel File')
    filePath2 = filePath.format(getpass.getuser())

def protectSheet():
    # global variable of excelSheetName
    global excelSheetName
    excelSheetName = str(entry_field.get())
    print(excelSheetName)

    wb1 = xl.load_workbook(filePath)
    ws1 = wb1[f'{excelSheetName}']

    newBook = Workbook()
    newSheet = newBook.active
    newSheet = newBook.create_sheet(f'{excelSheetName}', 0)
    del newBook['Sheet']

    # copy worksheet to another workbook
    max_rows = ws1.max_row
    max_columns = ws1.max_column

    for i in range(1, max_rows + 1):
        for j in range(1, max_columns + 1):
            #reading cell value from source excel file
            cellValue = ws1.cell(row = i, column = j)
            # writing cellValue to destination excel file
            newSheet.cell(row = i, column = j).value = cellValue.value

    # resize width of columns
    for col in newSheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        newSheet.column_dimensions[column].width = max_length

    redFill = PatternFill(start_color='db0000', end_color='db0000', fill_type='solid')
    greenFill = PatternFill(start_color='09D909', end_color='09D909', fill_type='solid')
    blueFill = PatternFill(start_color='2499FF', end_color='2499FF', fill_type='solid')
    pinkFill = PatternFill(start_color='FB80FF', end_color='FB80FF', fill_type='solid')
    greyFill = PatternFill(start_color='c4c4c4', end_color='c4c4c4', fill_type='solid')

    # color title row cells
    newSheet['A2'].fill = greyFill
    newSheet['B2'].fill = greyFill
    newSheet['C2'].fill = greyFill
    newSheet['D2'].fill = greyFill
    newSheet['E2'].fill = greyFill
    newSheet['F2'].fill = greyFill
    newSheet['G2'].fill = greyFill
    newSheet['H2'].fill = greyFill
    newSheet['I2'].fill = greyFill
    newSheet['J2'].fill = greyFill
    newSheet['K2'].fill = greyFill
    newSheet['L2'].fill = greyFill
    newSheet['M2'].fill = greyFill
    newSheet['N2'].fill = greyFill
    newSheet['O2'].fill = greyFill
    newSheet['P2'].fill = greyFill
    newSheet['Q2'].fill = greyFill

    # color cells before merging them
    newSheet['B1'].fill = redFill
    newSheet['C1'].fill = redFill
    newSheet['D1'].fill = redFill
    newSheet['E1'].fill = redFill

    newSheet['F1'].fill = greenFill
    newSheet['G1'].fill = greenFill
    newSheet['H1'].fill = greenFill
    newSheet['I1'].fill = greenFill
    newSheet['J1'].fill = greenFill

    newSheet['K1'].fill = blueFill
    newSheet['L1'].fill = blueFill
    newSheet['M1'].fill = blueFill

    newSheet['N1'].fill = pinkFill

    # merged cells row
    newSheet.merge_cells(start_row=1,start_column=2,end_row=1,end_column=5)
    newSheet.merge_cells(start_row=1,start_column=6,end_row=1,end_column=10)
    newSheet.merge_cells(start_row=1,start_column=11,end_row=1,end_column=13)

    newSheet['B1'].value = 'PICK UP'
    newSheet['F1'].value = 'DROP OFF'
    newSheet['K1'].value = 'PICK UP'
    newSheet['N1'].value = 'DROP OFF'

    newSheet.protection.password = '12345'
    finalPath = 'C:/Users/{}/' + excelSheetName + ' Invoice Sent.xlsx'
    finalFilePath = finalPath.format(getpass.getuser())
    newBook.save(filename=finalFilePath)

    print('Invoice Created.')

def protectSheet2():
    # global variable of excelSheetName
    global excelSheetName
    excelSheetName = str(entry_field.get())
    print(excelSheetName)

    wb1 = xl.load_workbook(filePath)
    ws1 = wb1[f'{excelSheetName}']

    newBook = Workbook()
    newSheet = newBook.active
    newSheet = newBook.create_sheet(f'{excelSheetName}', 0)
    del newBook['Sheet']

    # copy worksheet to another workbook
    max_rows = ws1.max_row
    max_columns = ws1.max_column

    for i in range(1, max_rows + 1):
        for j in range(1, max_columns + 1):
            # reading cell value from source excel file
            cellValue = ws1.cell(row=i, column=j)
            # writing cellValue to destination excel file
            newSheet.cell(row=i, column=j).value = cellValue.value

    # resize width of columns
    for col in newSheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        newSheet.column_dimensions[column].width = max_length

    greyFill = PatternFill(start_color='c4c4c4', end_color='c4c4c4', fill_type='solid')

    # color title row cells
    newSheet['A1'].fill = greyFill
    newSheet['B1'].fill = greyFill
    newSheet['C1'].fill = greyFill
    newSheet['D1'].fill = greyFill
    newSheet['E1'].fill = greyFill
    newSheet['F1'].fill = greyFill
    newSheet['G1'].fill = greyFill
    newSheet['H1'].fill = greyFill
    newSheet['I1'].fill = greyFill
    newSheet['J1'].fill = greyFill
    newSheet['K1'].fill = greyFill
    newSheet['L1'].fill = greyFill
    newSheet['M1'].fill = greyFill
    newSheet['N1'].fill = greyFill
    newSheet['O1'].fill = greyFill
    newSheet['P1'].fill = greyFill
    newSheet['Q1'].fill = greyFill

    # center alignment
    for i in range(1, 18):
        for j in range(1, 100):
            newSheet.cell(row=j, column=i).alignment = Alignment(horizontal='center')

    newSheet.protection.password = '12345'
    finalPath = 'C:/Users/{}/' + excelSheetName + ' Totes Invoice.xlsx'
    finalFilePath = finalPath.format(getpass.getuser())
    newBook.save(filename= finalFilePath)

    print('Totes Invoice Created.')


label_file_explorer = Label(root, text = 'Protect Worksheet of .xlsx File', width = 100, height = 4, background = 'black', foreground = 'white')
label_file_explorer.grid(column = 1, row = 1)
button_explore = Button(root, text = 'Select a .xlsx File', command = browseFile)
button_explore.grid(column = 1, row = 2, pady = 10)

label_input = Label(root, text = "Enter the month and year of invoice. e.g. January 2020:", width = 100, height = 4, background = 'black', foreground = 'white')
label_input.grid(column = 1, row = 3)

info = StringVar()
entry_field = Entry(root, textvariable = 'info')
entry_field.grid(column = 1, row = 4, pady = 10)
button_action = Button(root, text = 'Create/Protect the Invoice Worksheet', height=2, command = protectSheet)
button_action.grid(column = 1, row = 5, pady = 10)

button_action2 = Button(root, text = 'Create/Protect the Totes Worksheet', height=2, command = protectSheet2)
button_action2.grid(column = 1, row = 6, pady = 10)

exit_button = Button(root, text = 'Close', background = 'red', width = 10, height=2, command = root.destroy)
exit_button.grid(column = 1, row = 10, pady = 10)


root.mainloop()
